from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from os import path
import Setup

#To get chrome driver:
#sudo apt-get install chromium-chromedriver

directory = r'C:\Users\FireFly\PycharmProjects\pythonProject\filtered data'

def getDetails(coordinates):
    driver = webdriver.Chrome(executable_path= r'C:\Users\FireFly\PycharmProjects\pythonProject\driver\chromedriver.exe')
    # chrome_options = webdriver.ChromeOptions()
    #
    # # Starting webdriver with disable images option to save bandwidth.
    # prefs = {"profile.managed_default_content_settings.images": 2}
    # chrome_options.add_experimental_option("prefs", prefs)
    # chrome_options.add_argument('--disable-gpu')
    # chrome_options.add_argument('blink-settings=imagesEnabled=false')
    # #Running chrome without showing window (headless mode)
    # chrome_options.add_argument("headless") # Comment this line to stop headless mode (for debugging)
    #
    # driver = webdriver.Chrome(options=chrome_options)
    #
    actions = ActionChains(driver)

    driver.get("https://hsac.org.in/eodb/")  # Opening website

    #This function waits and checks if a dynamic element is present or not before continuing the script
    #If element is not found after a given timeout , it raises an exception.
    def checkElementPresence(timeout, xpath):
        try:
            element_present = EC.presence_of_element_located((By.XPATH, xpath))
            WebDriverWait(driver, timeout).until(element_present)
        except TimeoutException:
            print("Timed out waiting for" + str([{xpath[-10:]}]) + "to load")
            driver.quit()  # Quit the opened browser before ending the script.
            raise Exception("Time out error")

    checkElementPresence(
        20, '//*[@id="map"]/div[1]/div[3]/div[2]/div[2]/div/div/div[3]/form/input')

    #Find searchbox and enter coordinates.
    searchbox = driver.find_element_by_xpath(
        '//*[@id="map"]/div[1]/div[3]/div[2]/div[2]/div/div/div[3]/form/input')
    searchbox.send_keys(coordinates)
    #Find search button and click.
    searchbutton = driver.find_element_by_xpath(
        '//*[@id="map"]/div[1]/div[3]/div[2]/div[2]/div/div/div[4]')
    searchbutton.click()

    #Click on zoom button for accuracy
    checkElementPresence(
        20, '//*[@id="map"]/div[1]/div[3]/div[1]/div[2]/div[1]/div/div/div/span[1]')
    sleep(1)
    zoombutton = driver.find_element_by_xpath(
        '//*[@id="map"]/div[1]/div[3]/div[1]/div[2]/div[1]/div/div/div/span[1]')
    zoombutton.click()

    checkElementPresence(
        20, '//*[@id="map"]/div[1]/div[3]/div[1]/div[2]/div[2]/div')

    #Find the pointer pointing at the coordinates and click near it to get the details.
    pointer = driver.find_element_by_xpath(
        '//*[@id="map"]/div[1]/div[3]/div[1]/div[2]/div[2]/div')
    actions.move_to_element_with_offset(pointer, 13, 20)
    actions.click()
    actions.perform()

    checkElementPresence(
        20, '/html/body/div[1]/div[2]/div[1]/div[3]/div[1]/div[2]/div[1]/article/div/div/div/div/div/div[1]/table')

    #Reading details table
    data = driver.find_element_by_xpath(
        '/html/body/div[1]/div[2]/div[1]/div[3]/div[1]/div[2]/div[1]/article/div/div/div/div/div/div[1]/table')
    # get all of the rows in the table
    rows = data.find_elements(By.TAG_NAME, "tr")
    output_data = {}
    for row in rows:
        th = row.find_element(By.TAG_NAME, "th")
        td = row.find_element(By.TAG_NAME, "td")
        # Adding to output_data dictionary
        output_data.update({th.text: td.text})

    #Making a list for owners name.
    owner_list = []
    try:
        element_present = EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[3]/div[1]/div[2]/div[1]/article/div/div/div/div/div/div[2]/b/table/tbody'))
        # Wait for owners name to appear
        WebDriverWait(driver, 7).until(element_present)
        # The table containing owners name
        data = driver.find_element_by_xpath(
            '/html/body/div[1]/div[2]/div[1]/div[3]/div[1]/div[2]/div[1]/article/div/div/div/div/div/div[2]/b/table/tbody')
        # get all of the rows in the table
        rows = data.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            owner_list.append(row.text)
        # Adding to dictionary as a list.
        output_data.update({"Owners Name": owner_list})
    except TimeoutException:  # If timed out
        print("Owner List Empty")
        # Adding to dictionary as a list.
        output_data.update({"Owners Name": ['No Owner Names']})

    driver.quit()  # Closing the browser

    #Writing to text file
    # with open('result.txt', 'w') as f:
    #     print(output_data, file=f)
    return(output_data)

in_file = path.join(Setup.output_folder, 'fire.xlsx')
out_file = path.join(Setup.output_folder, 'farmer_detail.xlsx')

def locate():
    read_wb = load_workbook(filename=in_file)
    read_sheet = read_wb.active
    total_rows = read_sheet.max_row
    print("Total rows =" + str(total_rows))

    #If output excel file doesn't exist, create it.
    if(path.isfile(out_file) == False):
        write_wb = Workbook()
        write_wb.save(filename=out_file)

    #This will open the output excel file in append mode
    write_wb = load_workbook(filename=out_file)
    write_sheet = write_wb.active
    #
    last_filled_row = write_sheet.max_row  # Used for resuming loop

    #Setting auto sized column width for readability
    for i in range(1, 12):
        write_sheet.column_dimensions[get_column_letter(i)].auto_size = True

    #Column Headings
    write_sheet.cell(row=1, column=1).value = "State"
    write_sheet.cell(row=1, column=2).value = "District"
    write_sheet.cell(row=1, column=3).value = "Tehsil"
    write_sheet.cell(row=1, column=4).value = "Village Name"
    write_sheet.cell(row=1, column=5).value = "Longitude"
    write_sheet.cell(row=1, column=6).value = "Latitude"
    write_sheet.cell(row=1, column=9).value = "MurabbaNo"
    write_sheet.cell(row=1, column=10).value = "KhasaraNo"
    write_sheet.cell(row=1, column=11).value = "OwnerName"


    #loop records and search for coordinates
    for i in range(2, 6):
        Y = read_sheet.cell(row=i, column=1).value
        X = read_sheet.cell(row=i, column=2).value
        coordinates = str(Y) + ", " + str(X)
        print("Use " + "Coordinates =" + str(coordinates))
        result = None
        trycount = 0
        #Keep trying until you get result
        while result == None and trycount<5:
            try:
                result = getDetails(coordinates)
            except:
                trycount = trycount+1
                pass
        if(result==None):
            print("Network/server problem or invalid coordinates, please check manually before trying again.")
            exit()
        print(result, "\n")

        #Writing this row to output file
        write_sheet.cell(row=i, column=1).value = "Haryana"  # Same for all
        write_sheet.cell(row=i, column=2).value = result['District Name']
        write_sheet.cell(row=i, column=3).value = result['Tehsil Name']
        write_sheet.cell(row=i, column=4).value = result['Village Name']
        write_sheet.cell(row=i, column=5).value = X
        write_sheet.cell(row=i, column=6).value = Y
        write_sheet.cell(row=i, column=9).value = result['Murabba No']
        write_sheet.cell(row=i, column=10).value = result['Khasra No']

        #For owners list
        st_owners = ""
        for owner in result['Owners Name']:
            st_owners = st_owners+", "+owner
            write_sheet.cell(row=i, column=11).value = st_owners[1:]

        #Saving this row
        write_wb.save(filename=out_file)
